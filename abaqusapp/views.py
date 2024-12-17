from django.shortcuts import render
from rest_framework.views import APIView
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from .forms import ExcelUploadForm
import openpyxl
from .models import Portfolio, Asset, Weight, Price, InitialInvestment
from django.views.generic import ListView, DetailView
from rest_framework.response import Response
from django.utils import timezone

class PortfolioListView(ListView):
    model = Portfolio
    template_name = 'portfolio/list.html'
    context_object_name = 'portfolios'

class PortfolioDetailView(DetailView):
    model = Portfolio
    template_name = 'portfolio/detail.html'
    context_object_name = 'portfolio'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        portfolio = self.object
        
        # Call the PortfolioDataView API
        view = PortfolioDataView()
        request = self.request

        # Fix: Access the 'format' parameter and 'end_date' variable
        format = request.GET.get('format', None)
        if 'fecha_fin' in request.query_params:
            end_date = timezone.datetime.strptime(request.query_params['fecha_fin'], '%Y-%m-%d')
            context['end_date'] = end_date

        return context

class PortfolioDataView(APIView):
    def get(self, request, format=None):
        # Fix: Access the 'format' parameter
        format = request.GET.get('format', None)
        # Your logic here
        return Response({'message': 'Data retrieved successfully'})

class ExcelUploadView(APIView):
    def get(self, request):
        form = ExcelUploadForm()
        return render(request, 'portfolio/upload.html', {'form': form})

    def post(self, request):
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)

            # Verificar si las hojas existen
            if 'precios' not in wb.sheetnames:
                return HttpResponse("La hoja 'precios' no existe en el archivo Excel.", status=400)
            if 'weights' not in wb.sheetnames:
                return HttpResponse("La hoja 'weights' no existe en el archivo Excel.", status=400)

            # Procesar la hoja de precios
            sheet = wb['precios']
            assets = ["EEUU", "Europa", "Japón", "EM Asia", "Latam", "High Yield", "IG Corporate", "EMHC", "Latam HY", "UK", "Asia Desarrollada", "EMEA", "Otros RV", "Tesoro", "MBS+CMBS+AMBS", "ABS", "MM/Caja"]
            asset_objects = {name: Asset.objects.get_or_create(name=name)[0] for name in assets}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                date, *values = row
                for asset, value in zip(assets, values):
                    Price.objects.create(date=date, asset=asset_objects[asset], price=value)

            # Procesar la hoja de pesos
            sheet = wb['weights']
            portfolios = {1: Portfolio.objects.get_or_create(name='Portfolio 1', initial_value=1000000)[0],
                          2: Portfolio.objects.get_or_create(name='Portfolio 2', initial_value=1000000)[0]}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                date, asset_name, portfolio_1_weight, portfolio_2_weight = row
                asset = asset_objects[asset_name]
                Weight.objects.create(date=date, asset=asset, portfolio=portfolios[1], weight=portfolio_1_weight)
                Weight.objects.create(date=date, asset=asset, portfolio=portfolios[2], weight=portfolio_2_weight)

            # Calcular las cantidades iniciales invertidas
            for portfolio in portfolios.values():
                initial_value = portfolio.initial_value
                for weight in Weight.objects.filter(portfolio=portfolio, date='2022-02-15'):
                    price = Price.objects.get(asset=weight.asset, date='2022-02-15').price
                    quantity = (weight.weight * initial_value) / price
                    InitialInvestment.objects.create(asset=weight.asset, portfolio=portfolio, quantity=quantity)

            return HttpResponseRedirect(reverse('portfolio_list'))
        return render(request, 'portfolio/upload.html', {'form': form})
    
def index(request):
    return render(request, 'index.html')
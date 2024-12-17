import openpyxl
from .models import Asset, Portfolio, PortfolioAllocation

def load_data_from_excel():
    workbook = openpyxl.load_workbook('data.xlsx')

    # Cargar datos de la hoja "Weights"
    weights_sheet = workbook['Weights']
    for row in range(2, weights_sheet.max_row + 1):
        portfolio_1_weight = weights_sheet.cell(row=row, column=3).value
        portfolio_2_weight = weights_sheet.cell(row=row, column=4).value
        asset_name = weights_sheet.cell(row=row, column=2).value

        asset, _ = Asset.objects.get_or_create(name=asset_name)

        portfolio_1, _ = Portfolio.objects.get_or_create(name='Portfolio 1')
        portfolio_1.weights[asset_name] = portfolio_1_weight
        portfolio_1.save()

        portfolio_2, _ = Portfolio.objects.get_or_create(name='Portfolio 2')
        portfolio_2.weights[asset_name] = portfolio_2_weight
        portfolio_2.save()

        PortfolioAllocation.objects.get_or_create(
            portfolio=portfolio_1,
            asset=asset,
            quantity=0
        )
        PortfolioAllocation.objects.get_or_create(
            portfolio=portfolio_2,
            asset=asset,
            quantity=0
        )

    # Cargar datos de la hoja "Precios"
    prices_sheet = workbook['Precios']
    for row in range(2, prices_sheet.max_row + 1):
        date = prices_sheet.cell(row=row, column=1).value.date()
        for col in range(2, prices_sheet.max_column + 1):
            asset_name = prices_sheet.cell(row=1, column=col).value
            asset, _ = Asset.objects.get_or_create(name=asset_name)
            asset.price_history[str(date)] = prices_sheet.cell(row=row, column=col).value
            asset.save()

def calculate_initial_quantities():
    for portfolio in Portfolio.objects.all():
        for allocation in PortfolioAllocation.objects.filter(portfolio=portfolio):
            asset = allocation.asset
            weight = portfolio.weights[asset.name]
            allocation.quantity = (weight * portfolio.initial_value) / asset.price_history[str(portfolio.created_at.date())]
            allocation.save()